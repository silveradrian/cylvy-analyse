#!/usr/bin/env python3
"""
Database Manager Module

This module provides a SQLite database interface for storing analysis jobs,
results, and other application data. It handles database initialization,
queries, and data export functionality.
"""

import os
import sqlite3
import json
import time
import csv
import logging
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple, Union

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger("db_manager")

# Default database location
DEFAULT_DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "analysis.db")

class DatabaseManager:
    """
    Manages database operations for the content analysis application.
    """
    
    def __init__(self, db_path: str = None):
        """
        Initialize the database manager.
        
        Args:
            db_path: Path to the SQLite database file
        """
        if not db_path:
            # Default database location in data directory
            data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')
            os.makedirs(data_dir, exist_ok=True)
            db_path = os.path.join(data_dir, 'analysis.db')
                
        self.db_path = db_path
        self.init_db()  # Changed from self._init_db() to self.init_db()
        self.db = self  # Add this line - this will make self.db.method() work by referring to self
        logger.info(f"Database initialized at {db_path}")
    
    def get_connection(self) -> sqlite3.Connection:
        """
        Get a database connection with row factory for dict-like access.
        
        Returns:
            A SQLite connection object
        """
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn
    
    def init_db(self) -> None:
        """
        Initialize the database schema if it doesn't exist.
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Create jobs table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS jobs (
            job_id TEXT PRIMARY KEY,
            name TEXT,
            status TEXT,
            created_at REAL,
            updated_at REAL,
            completed_at REAL,
            total_urls INTEGER,
            processed_urls INTEGER,
            error_count INTEGER,
            prompt_names TEXT,
            company_info TEXT
        )
        ''')
        
        # Create results table for storing analysis results
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id TEXT,
            url TEXT,
            title TEXT,
            status TEXT,
            content_type TEXT,
            word_count INTEGER,
            processed_at REAL,
            prompt_name TEXT,
            api_tokens INTEGER,
            error TEXT,
            data TEXT,
            FOREIGN KEY (job_id) REFERENCES jobs (job_id)
        )
        ''')
        
        # Create prompt_usage table to track prompt usage
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS prompt_usage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            prompt_name TEXT,
            job_id TEXT,
            url TEXT,
            tokens_used INTEGER,
            processed_at REAL,
            success BOOLEAN,
            FOREIGN KEY (job_id) REFERENCES jobs (job_id)
        )
        ''')
        
        # Create indexes for performance
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_results_job_id ON results (job_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_results_url ON results (url)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_prompt_usage_job_id ON prompt_usage (job_id)')
        
        conn.commit()
        conn.close()
    
    def create_job(self, urls: List[str], prompts: List[str], name: str = None, company_info: Dict = None) -> str:
        """
        Create a new job in the database.
        
        Args:
            urls: List of URLs to analyze
            prompts: List of prompt names to use
            name: Optional job name
            company_info: Optional company information dictionary
                
        Returns:
            The job ID
        """
        try:
            import uuid as uuid_module  # Import locally to avoid namespace issues
            job_id = str(uuid_module.uuid4())
            current_time = time.time()
            
            if name is None:
                name = f"Analysis {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                
            if company_info is None:
                company_info = {}
            
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute(
                '''
                INSERT INTO jobs (
                    job_id, name, status, created_at, updated_at, 
                    total_urls, processed_urls, error_count, prompt_names, company_info
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    job_id,
                    name,
                    'pending',
                    current_time,
                    current_time,
                    len(urls),
                    0,
                    0,
                    json.dumps(prompts),
                    json.dumps(company_info)
                )
            )
            
            conn.commit()
            conn.close()
            
            logger.info(f"Created new job with ID: {job_id}")
            return job_id
            
        except Exception as e:
            logger.error(f"Error creating job {name}: {str(e)}")
            return None

    # Still return job_id for backward compatibility
    
    def update_job_status(self, job_id: str, status: str = None, 
                     processed_urls: Optional[int] = None,
                     total_urls: Optional[int] = None,
                     error_count: Optional[int] = None,
                     completed_at: Optional[Union[float, str]] = None) -> bool:
        """
        Update the status and counters of a job.
        
        Args:
            job_id: The job identifier
            status: New status value
            processed_urls: Number of processed URLs
            total_urls: Total number of URLs
            error_count: Number of errors
            completed_at: Timestamp when job completed
            
        Returns:
            True if successful, False otherwise
        """
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Get current job data
            cursor.execute('SELECT * FROM jobs WHERE job_id = ?', (job_id,))
            job = cursor.fetchone()
            
            if not job:
                logger.warning(f"Job {job_id} not found when updating status")
                return False
            
            update_fields = ['updated_at = ?']
            params = [time.time()]
            
            if status is not None:
                update_fields.append('status = ?')
                params.append(status)
                    
            if processed_urls is not None:
                update_fields.append('processed_urls = ?')
                params.append(processed_urls)
                    
            if total_urls is not None:
                update_fields.append('total_urls = ?')
                params.append(total_urls)
                    
            if error_count is not None:
                update_fields.append('error_count = ?')
                params.append(error_count)
                    
            # If completed_at is provided or status is 'completed'/'completed_with_errors'/'failed'
            if completed_at is not None:
                update_fields.append('completed_at = ?')
                # Handle string timestamp or use the value directly
                if isinstance(completed_at, str):
                    try:
                        dt = datetime.fromisoformat(completed_at.replace('Z', '+00:00'))
                        params.append(dt.timestamp())
                    except ValueError:
                        params.append(time.time())
                else:
                    params.append(completed_at)
            elif status in ['completed', 'completed_with_errors', 'failed']:
                update_fields.append('completed_at = ?')
                params.append(time.time())
            
            # Build the update query
            query = f"UPDATE jobs SET {', '.join(update_fields)} WHERE job_id = ?"
            params.append(job_id)
            
            cursor.execute(query, params)
            conn.commit()
            conn.close()
            
            logger.info(f"Updated job {job_id} status to {status}")
            return True
                
        except Exception as e:
            logger.error(f"Error updating job {job_id} status: {str(e)}")
            return False
    

    def save_result(self, result):
        """
        Save an analysis result to the database.
        
        Args:
            result: Dictionary containing the analysis result
                
        Returns:
            The ID of the saved result
        """
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Extract values from result dictionary with fallbacks
            job_id = result.get('job_id')
            url = result.get('url', '')
            title = result.get('title', '')
            status = result.get('status', 'unknown')
            content_type = result.get('content_type', 'unknown')
            word_count = int(result.get('word_count', 0) or 0)  # Convert None to 0
            api_tokens = int(result.get('api_tokens', 0) or 0)  # Convert None to 0
            error = result.get('error', '')
            prompt_name = result.get('prompt_name', '')
            
            # Log basic info about the result being saved
            logger.info(f"Saving result for URL: {url}, job_id: {job_id}")
            
            # Initialize data object to store all structured data
            data_obj = {}
            
            # 1. Include analysis_results if present
            if 'analysis_results' in result and result['analysis_results']:
                data_obj['analysis_results'] = result['analysis_results']
                analysis_keys = list(result['analysis_results'].keys())
                logger.info(f"Including analysis_results with {len(analysis_keys)} prompts: {analysis_keys}")
            
            # 2. Include structured_data if present
            if 'structured_data' in result and result['structured_data']:
                data_obj['structured_data'] = result['structured_data']
                # Calculate total fields across all prompts
                total_fields = sum(
                    len(fields) for prompt, fields in result['structured_data'].items()
                    if isinstance(fields, dict)
                )
                prompt_counts = {
                    prompt: len(fields) for prompt, fields in result['structured_data'].items()
                    if isinstance(fields, dict)
                }
                logger.info(f"Including structured_data with {total_fields} total fields across prompts: {prompt_counts}")
            
            # 3. If data is already a serialized string, parse and incorporate it
            if 'data' in result and result['data']:
                # Handle string JSON data
                if isinstance(result['data'], str) and result['data'].strip() not in ('{}', ''):
                    try:
                        existing_data = json.loads(result['data']) 
                        if isinstance(existing_data, dict):
                            # Merge with our data object, preferring any data we've already collected
                            for key, value in existing_data.items():
                                if key not in data_obj:
                                    data_obj[key] = value
                                elif key == 'structured_data' and isinstance(value, dict):
                                    # For structured_data, merge at the prompt level
                                    for prompt, fields in value.items():
                                        if prompt not in data_obj['structured_data']:
                                            data_obj['structured_data'][prompt] = fields
                            logger.info(f"Merged existing data with keys: {list(existing_data.keys())}")
                    except json.JSONDecodeError:
                        logger.warning(f"Could not parse existing data JSON: {result['data'][:100]}...")
                # Handle dictionary data
                elif isinstance(result['data'], dict):
                    # Similar merge logic for dict data
                    for key, value in result['data'].items():
                        if key not in data_obj:
                            data_obj[key] = value
                        elif key == 'structured_data' and isinstance(value, dict):
                            if 'structured_data' not in data_obj:
                                data_obj['structured_data'] = {}
                            for prompt, fields in value.items():
                                if prompt not in data_obj['structured_data']:
                                    data_obj['structured_data'][prompt] = fields
                    logger.info(f"Merged data dict with keys: {list(result['data'].keys())}")
            
            # 4. Extract direct structured fields from the result
            structured_fields = {}
            for key, value in result.items():
                # Identify fields that look like structured data (prefixed or containing underscore)
                if (key.startswith(('ci_', 'pa_', 'ca_')) or 
                    ('_' in key and key not in ('job_id', 'api_tokens', 'word_count', 'content_type'))):
                    structured_fields[key] = value
            
            # If we found direct structured fields, include them
            if structured_fields:
                # Ensure structured_data container exists
                if 'structured_data' not in data_obj:
                    data_obj['structured_data'] = {}
                
                # Use the prompt name if available, or default
                target_prompt = prompt_name or 'content_analysis'
                if target_prompt not in data_obj['structured_data']:
                    data_obj['structured_data'][target_prompt] = {}
                
                # Add all structured fields
                data_obj['structured_data'][target_prompt].update(structured_fields)
                logger.info(f"Added {len(structured_fields)} direct structured fields to target prompt '{target_prompt}'")

            # Convert the final data object to JSON
            data_json = json.dumps(data_obj)
            
            # Log data size and structure
            logger.info(f"Final data JSON size: {len(data_json)} chars, top-level keys: {list(data_obj.keys())}")
            
            # Use processed_at instead of created_at for the timestamp
            processed_at = result.get('processed_at', time.time())
            
            # Insert the result
            cursor.execute(
                """
                INSERT INTO results (
                    job_id, url, title, status, content_type, word_count, 
                    processed_at, prompt_name, api_tokens, error, data
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (job_id, url, title, status, content_type, word_count, 
                 processed_at, prompt_name, api_tokens, error, data_json)
            )
            
            result_id = cursor.lastrowid
            logger.info(f"Saved result to database with ID: {result_id}")
            
            # If there are API tokens used, also log in prompt_usage table
            if api_tokens > 0 and prompt_name:
                try:
                    cursor.execute(
                        """
                        INSERT INTO prompt_usage (
                            prompt_name, job_id, url, tokens_used, processed_at, success
                        )
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (prompt_name, job_id, url, api_tokens, processed_at, status == 'success')
                    )
                    logger.info(f"Logged token usage: {api_tokens} tokens for prompt '{prompt_name}'")
                except Exception as e:
                    logger.warning(f"Failed to insert prompt usage: {e}")
            
            conn.commit()
            conn.close()
            return result_id
            
        except sqlite3.Error as e:
            logger.error(f"Database error in save_result: {e}")
            return None
        except Exception as e:
            logger.error(f"Error saving result: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return None


    def get_job(self, job_id: str) -> Optional[Dict[str, Any]]:
        """
        Get job details by ID.
        
        Args:
            job_id: The job identifier
            
        Returns:
            Job details or None if not found
        """
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM jobs WHERE job_id = ?', (job_id,))
            job = cursor.fetchone()
            
            if not job:
                return None
            
            # Convert to dict
            job_dict = dict(job)
            
            # Parse JSON fields
            if 'prompt_names' in job_dict and job_dict['prompt_names']:
                job_dict['prompt_names'] = json.loads(job_dict['prompt_names'])
            else:
                job_dict['prompt_names'] = []
                
            if 'company_info' in job_dict and job_dict['company_info']:
                job_dict['company_info'] = json.loads(job_dict['company_info'])
            else:
                job_dict['company_info'] = {}
            
            # Calculate progress percentage
            total = job_dict.get('total_urls', 0)
            processed = job_dict.get('processed_urls', 0)
            
            if total > 0:
                job_dict['progress'] = (processed / total) * 100
            else:
                job_dict['progress'] = 0
            
            # Add formatted timestamps
            for ts_field in ['created_at', 'updated_at', 'completed_at']:
                if job_dict.get(ts_field):
                    job_dict[f"{ts_field}_formatted"] = datetime.fromtimestamp(
                        job_dict[ts_field]).strftime('%Y-%m-%d %H:%M:%S')
            
            conn.close()
            return job_dict
            
        except Exception as e:
            logger.error(f"Error getting job {job_id}: {str(e)}")
            return None
    
    
    def get_results_for_job(self, job_id: str) -> List[Dict[str, Any]]:
        """
        Get all results for a specific job with structured data properly extracted.
        
        Args:
            job_id: The job ID
            
        Returns:
            List of result dictionaries with parsed data
        """
        try:
            conn = self.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("SELECT * FROM results WHERE job_id = ? ORDER BY processed_at DESC", (job_id,))
            rows = cursor.fetchall()
            
            # Log the first row structure for debugging
            if rows:
                first_row = dict(rows[0])
                logger.info(f"Raw result keys: {list(first_row.keys())}")
                if 'data' in first_row:
                    logger.info(f"Data type: {type(first_row['data'])}")
                    if isinstance(first_row['data'], str):
                        logger.info(f"First 100 chars of data: {first_row['data'][:100]}")
            else:
                logger.warning(f"No results found for job {job_id}")
                return []
            
            # Convert rows to dicts and parse data
            results = []
            for i, row in enumerate(rows):
                result = dict(row)
                
                # Parse data field if present
                if 'data' in result and result['data']:
                    try:
                        # If data is a string, try to parse it as JSON
                        if isinstance(result['data'], str):
                            parsed_data = json.loads(result['data'])
                            result['data'] = parsed_data
                            
                            # Debug only first result to avoid log spam
                            if i == 0:
                                logger.info(f"Parsed data keys: {list(parsed_data.keys()) if isinstance(parsed_data, dict) else 'Not a dict'}")
                        
                        # If parsing succeeds or data was already a dict
                        if isinstance(result['data'], dict):
                            data = result['data']
                            
                            # Extract structured fields directly into result for easy access
                            # From analysis_results
                            if 'analysis_results' in data:
                                for prompt_name, analysis in data['analysis_results'].items():
                                    if isinstance(analysis, dict) and 'parsed_fields' in analysis:
                                        if i == 0:  # Debug only first result
                                            logger.info(f"Found parsed_fields in prompt {prompt_name}: {list(analysis['parsed_fields'].keys())}")
                                        for field_name, field_value in analysis['parsed_fields'].items():
                                            result[field_name] = field_value
                            
                            # From structured_data
                            if 'structured_data' in data:
                                for prompt_name, fields in data['structured_data'].items():
                                    if isinstance(fields, dict):
                                        if i == 0:  # Debug only first result
                                            logger.info(f"Found structured_data in prompt {prompt_name}: {list(fields.keys())}")
                                        for field_name, field_value in fields.items():
                                            result[field_name] = field_value
                    
                    except (json.JSONDecodeError, TypeError) as e:
                        logger.warning(f"Failed to parse data JSON for URL {result.get('url')}: {str(e)}")
                
                results.append(result)
            
            # Debug summary
            if results:
                fields_found = set()
                for r in results:
                    for k in r.keys():
                        if k not in ['id', 'job_id', 'url', 'title', 'status', 'content_type', 
                                    'word_count', 'processed_at', 'prompt_name', 'api_tokens', 
                                    'error', 'data']:
                            fields_found.add(k)
                
                logger.info(f"Found {len(fields_found)} structured fields in results: {sorted(list(fields_found))}")
            
            conn.close()
            return results
        
        except Exception as e:
            logger.error(f"Error in get_results_for_job: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return []
    
    def get_all_jobs(self, limit: int = 100, offset: int = 0) -> List[Dict[str, Any]]:
        """
        Retrieve all jobs with pagination, ordered by creation date (most recent first).
        
        Args:
            limit: Maximum number of jobs to return
            offset: Number of jobs to skip
                
        Returns:
            List of job dictionaries
        """
        try:
            # Ensure limit and offset are integers
            try:
                limit = int(limit) if limit is not None else 100
                offset = int(offset) if offset is not None else 0
            except (ValueError, TypeError):
                logger.warning(f"Invalid parameters: limit={limit}, offset={offset}. Using defaults.")
                limit = 100
                offset = 0
                
            # Ensure values are positive
            limit = max(1, limit)
            offset = max(0, offset)
            
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Use parameterized query with explicit integer values
            cursor.execute(
                """
                SELECT job_id, name, status, created_at, updated_at, completed_at, 
                       total_urls, processed_urls, error_count, prompt_names, company_info
                FROM jobs
                ORDER BY created_at DESC
                LIMIT ? OFFSET ?
                """, 
                (int(limit), int(offset))  # Explicit casting to integers
            )
            
            rows = cursor.fetchall()
            cols = [col[0] for col in cursor.description]
            jobs = []
            
            for row in rows:
                # Create dictionary from row data
                job_data = dict(zip(cols, row))
                
                # Parse JSON fields safely
                try:
                    prompt_names = json.loads(job_data['prompt_names']) if job_data['prompt_names'] else []
                except (TypeError, json.JSONDecodeError):
                    prompt_names = []
                    
                try:
                    company_info = json.loads(job_data['company_info']) if job_data['company_info'] else {}
                except (TypeError, json.JSONDecodeError):
                    company_info = {}
                
                # Format timestamps safely
                try:
                    created_at = datetime.fromtimestamp(job_data['created_at']).isoformat() if job_data['created_at'] else None
                except (TypeError, ValueError):
                    created_at = None
                    
                try:
                    updated_at = datetime.fromtimestamp(job_data['updated_at']).isoformat() if job_data['updated_at'] else None
                except (TypeError, ValueError):
                    updated_at = None
                    
                try:
                    completed_at = datetime.fromtimestamp(job_data['completed_at']).isoformat() if job_data['completed_at'] else None
                except (TypeError, ValueError):
                    completed_at = None
                
                # Create formatted job object
                job = {
                    "job_id": job_data['job_id'],
                    "name": job_data['name'] or "",
                    "status": job_data['status'] or "unknown",
                    "created_at": created_at,
                    "updated_at": updated_at,
                    "completed_at": completed_at,
                    "total_urls": job_data['total_urls'] or 0,
                    "processed_urls": job_data['processed_urls'] or 0,
                    "error_count": job_data['error_count'] or 0,
                    "prompts": prompt_names,
                    "prompt_count": len(prompt_names),
                    "company_info": company_info,
                    "progress_percentage": 0
                }
                
                # Calculate progress percentage
                if job["total_urls"] > 0:
                    job["progress_percentage"] = int((job["processed_urls"] / job["total_urls"]) * 100)
                
                # For backward compatibility
                job["url_count"] = job["total_urls"]
                
                jobs.append(job)
            
            conn.close()
            return jobs
                
        except sqlite3.Error as e:
            logger.error(f"Database error in get_all_jobs: {e}")
            return []
        except Exception as e:
            logger.error(f"Unexpected error in get_all_jobs: {str(e)}")
            return []
    
    def get_job_metrics(self, job_id: str) -> Dict[str, Any]:
        """
        Get metrics for a specific job.
        
        Args:
            job_id: The job identifier
            
        Returns:
            Dictionary of job metrics
        """
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Get job basics
            job = self.get_job(job_id)
            
            if not job:
                return {"error": "Job not found"}
            
            # Get result stats
            cursor.execute('''
            SELECT 
                COUNT(*) as total_results,
                SUM(CASE WHEN status = 'success' THEN 1 ELSE 0 END) as successful,
                SUM(CASE WHEN status = 'error' THEN 1 ELSE 0 END) as failed,
                SUM(word_count) as total_words,
                SUM(api_tokens) as total_tokens
            FROM results 
            WHERE job_id = ?
            ''', (job_id,))
            
            result_stats = dict(cursor.fetchone())
            
            # Get prompt usage stats
            cursor.execute('''
            SELECT 
                prompt_name,
                COUNT(*) as usage_count,
                SUM(tokens_used) as tokens_used 
            FROM prompt_usage 
            WHERE job_id = ? 
            GROUP BY prompt_name
            ''', (job_id,))
            
            prompt_usage = []
            for row in cursor.fetchall():
                prompt_usage.append(dict(row))
            
            # Calculate time taken if job is completed
            time_taken = None
            if job.get('completed_at') and job.get('created_at'):
                time_taken = job['completed_at'] - job['created_at']
            
            metrics = {
                "job_id": job_id,
                "status": job.get('status'),
                "total_urls": job.get('total_urls', 0),
                "processed_urls": job.get('processed_urls', 0),
                "progress": job.get('progress', 0),
                "total_results": result_stats.get('total_results', 0),
                "successful_results": result_stats.get('successful', 0),
                "failed_results": result_stats.get('failed', 0),
                "total_words": result_stats.get('total_words', 0),
                "total_tokens": result_stats.get('total_tokens', 0),
                "prompt_usage": prompt_usage,
                "time_taken": time_taken,
                "time_taken_formatted": self._format_duration(time_taken) if time_taken else None
            }
            
            conn.close()
            return metrics
            
        except Exception as e:
            logger.error(f"Error getting metrics for job {job_id}: {str(e)}")
            return {"error": str(e)}
    
    def export_results_to_csv(self, job_id):
        """Export job results to CSV with all structured fields, handling double-JSON encoding."""
        try:
            import tempfile, os, json, csv, logging
            
            logger = logging.getLogger("db_manager")
            
            # Get results for the job
            results = self.get_results_for_job(job_id)
            
            if not results:
                logger.warning(f"No results found for job {job_id}")
                return ""
                
            # Create a temporary file
            fd, csv_path = tempfile.mkstemp(suffix='.csv')
            os.close(fd)
            
            # Collect all fields - start with standard fields
            all_fields = set(['url', 'status', 'title', 'word_count', 'processed_at', 
                             'content_type', 'prompt_name', 'api_tokens', 'error'])
            
            # Helper function to extract structured data from analysis text
            def extract_structured_fields(analysis_text):
                fields = {}
                if not analysis_text:
                    return fields
                    
                # Process line by line
                for line in analysis_text.split('\n'):
                    if '|||' in line:
                        parts = line.split('|||', 1)  # Split only on first occurrence
                        if len(parts) == 2:
                            field_name = parts[0].strip()
                            field_value = parts[1].strip()
                            if field_name and field_value:
                                fields[field_name] = field_value
                return fields
                
            # First pass: collect all field names from the structured data
            for result in results:
                # First check for direct structured fields in the result (top level)
                for key in result.keys():
                    if key not in all_fields and key != 'data':
                        all_fields.add(key)
                
                # Then check in the data object
                if 'data' in result and result['data']:
                    data_obj = None
                    
                    # Try to parse the data - could be string or already parsed
                    if isinstance(result['data'], str):
                        try:
                            data_obj = json.loads(result['data'])
                        except json.JSONDecodeError:
                            logger.warning(f"Could not parse data JSON for result")
                            continue
                    else:
                        data_obj = result['data']
                        
                    if not isinstance(data_obj, dict):
                        continue
                    
                    # Look for structured_data
                    if 'structured_data' in data_obj and isinstance(data_obj['structured_data'], dict):
                        for prompt_name, fields in data_obj['structured_data'].items():
                            if isinstance(fields, dict):
                                for field_name in fields:
                                    all_fields.add(field_name)
                    
                    # Process analysis results for fields
                    if 'analysis_results' in data_obj and isinstance(data_obj['analysis_results'], dict):
                        for prompt_name, analysis in data_obj['analysis_results'].items():
                            if isinstance(analysis, dict):
                                # Check for direct structured data
                                if 'parsed_fields' in analysis and isinstance(analysis['parsed_fields'], dict):
                                    for field_name in analysis['parsed_fields']:
                                        all_fields.add(field_name)
                                        
                                # Process the analysis text for ||| delimited fields
                                if 'analysis' in analysis and analysis['analysis']:
                                    structured_fields = extract_structured_fields(analysis['analysis'])
                                    for field_name in structured_fields:
                                        all_fields.add(field_name)
            
            # Log found field count
            logger.info(f"Found {len(all_fields)} total fields to export")
            
            # Second pass: Extract all values
            rows = []
            for result in results:
                # Start with basic fields
                row = {field: '' for field in all_fields}
                
                # Fill in standard fields first
                for field in ['url', 'status', 'title', 'word_count', 'processed_at', 
                             'content_type', 'prompt_name', 'api_tokens', 'error']:
                    if field in result:
                        # Format timestamp
                        if field == 'processed_at' and result[field]:
                            try:
                                row[field] = datetime.fromtimestamp(result[field]).strftime('%Y-%m-%d %H:%M:%S')
                            except (ValueError, TypeError):
                                row[field] = result[field]
                        else:
                            row[field] = result[field]
                
                # Add top-level fields (direct structured fields)
                for key in result.keys():
                    if key in all_fields and key not in ['url', 'status', 'title', 'word_count', 'processed_at', 
                                                        'content_type', 'prompt_name', 'api_tokens', 'error', 'data']:
                        row[key] = result[key]
                
                # Now extract structured data from data object
                if 'data' in result and result['data']:
                    data_obj = None
                    
                    # Parse the data object
                    if isinstance(result['data'], str):
                        try:
                            data_obj = json.loads(result['data'])
                        except json.JSONDecodeError:
                            continue
                    else:
                        data_obj = result['data']
                    
                    if not isinstance(data_obj, dict):
                        continue
                    
                    # Extract from structured_data
                    if 'structured_data' in data_obj and isinstance(data_obj['structured_data'], dict):
                        for prompt_name, fields in data_obj['structured_data'].items():
                            if isinstance(fields, dict):
                                for field_name, value in fields.items():
                                    if field_name in all_fields:
                                        if isinstance(value, list):
                                            row[field_name] = ', '.join(str(v) for v in value if v is not None)
                                        else:
                                            row[field_name] = value
                    
                    # Extract from analysis results
                    if 'analysis_results' in data_obj and isinstance(data_obj['analysis_results'], dict):
                        for prompt_name, analysis in data_obj['analysis_results'].items():
                            if isinstance(analysis, dict):
                                # Extract from parsed_fields
                                if 'parsed_fields' in analysis and isinstance(analysis['parsed_fields'], dict):
                                    for field_name, value in analysis['parsed_fields'].items():
                                        if field_name in all_fields:
                                            if isinstance(value, list):
                                                row[field_name] = ', '.join(str(v) for v in value if v is not None)
                                            else:
                                                row[field_name] = value
                                
                                # Extract from analysis text with ||| delimiter
                                if 'analysis' in analysis and analysis['analysis']:
                                    structured_fields = extract_structured_fields(analysis['analysis'])
                                    for field_name, value in structured_fields.items():
                                        if field_name in all_fields:
                                            row[field_name] = value
                
                # Add the row to our results
                rows.append(row)
            
            # Find fields with actual values (skip empty fields)
            non_empty_fields = set()
            for row in rows:
                for field, value in row.items():
                    if value not in ('', None):
                        non_empty_fields.add(field)
            
            # Always include standard fields
            for field in ['url', 'status', 'title', 'word_count', 'processed_at', 
                         'content_type', 'prompt_name', 'api_tokens', 'error']:
                non_empty_fields.add(field)
                
            # Sort fields alphabetically
            sorted_fields = sorted(list(non_empty_fields))
                
            # Write to CSV file, but only include fields that have values
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=sorted_fields)
                writer.writeheader()
                
                # Write only the fields with values
                for row in rows:
                    filtered_row = {k: row[k] for k in sorted_fields}
                    writer.writerow(filtered_row)
            
            logger.info(f"Successfully exported to CSV: {csv_path}")
            return csv_path
        except Exception as e:
            logger.error(f"Error in CSV export: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return ""

    def export_results_to_excel(self, job_id):
        """Export job results to Excel with all structured fields, handling double-JSON encoding."""
        try:
            import tempfile, os, json, pandas as pd, logging
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            logger = logging.getLogger("db_manager")
            
            # First export to CSV to reuse the extraction logic
            csv_path = self.export_results_to_csv(job_id)
            
            if not csv_path or not os.path.exists(csv_path):
                logger.error(f"CSV export failed for job {job_id}")
                return ""
                
            # Create a temporary file for Excel
            fd, excel_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(fd)
            
            # Read the CSV into a pandas DataFrame
            df = pd.read_csv(csv_path)
            
            # Categorize fields for different sheets
            standard_fields = ['url', 'title', 'status', 'content_type', 'word_count', 
                             'processed_at', 'prompt_name', 'api_tokens', 'error']
            
            ci_fields = [col for col in df.columns if col.startswith('ci_')]
            pa_fields = [col for col in df.columns if col.startswith('pa_')]
            other_fields = [col for col in df.columns if col not in standard_fields + ci_fields + pa_fields]
            
            # Create Excel writer
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Main sheet with all data
                df.to_excel(writer, sheet_name='All Results', index=False)
                
                # Content Intelligence sheet if CI fields exist
                if ci_fields:
                    ci_df = df[['url', 'title'] + ci_fields]
                    ci_df.to_excel(writer, sheet_name='Content Intelligence', index=False)
                
                # Persona Analysis sheet if PA fields exist
                if pa_fields:
                    pa_df = df[['url', 'title'] + pa_fields]
                    pa_df.to_excel(writer, sheet_name='Persona Analysis', index=False)
                
                # Get job info for summary sheet
                job = self.get_job(job_id)
                metrics = self.get_job_metrics(job_id)
                
                # Create summary sheet
                summary_data = {
                    'Metric': [
                        'Job ID', 'Job Name', 'Status', 'Total URLs', 'Processed URLs',
                        'Success Rate', 'Total Words', 'Total API Tokens',
                        'Created At', 'Completed At'
                    ],
                    'Value': [
                        job_id,
                        job.get('name', ''),
                        job.get('status', ''),
                        job.get('total_urls', 0),
                        job.get('processed_urls', 0),
                        f"{(metrics.get('successful_results', 0) / max(metrics.get('total_results', 1), 1)) * 100:.1f}%",
                        metrics.get('total_words', 0),
                        metrics.get('total_tokens', 0),
                        job.get('created_at', ''),
                        job.get('completed_at', '')
                    ]
                }
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Adjust column widths after saving
            try:
                wb = load_workbook(excel_path)
                for sheet in wb.worksheets:
                    for col in sheet.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    cell_length = len(str(cell.value))
                                    max_length = max(max_length, cell_length)
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50
                        sheet.column_dimensions[column].width = adjusted_width
                wb.save(excel_path)
            except Exception as e:
                logger.warning(f"Failed to adjust column widths: {e}")
            
            # Clean up CSV file
            try:
                os.remove(csv_path)
            except:
                pass
            
            logger.info(f"Successfully exported to Excel: {excel_path}")
            return excel_path
        except Exception as e:
            logger.error(f"Error in Excel export: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return ""
    
    def _format_duration(self, seconds: float) -> str:
        """Format a duration in seconds to a human-readable string."""
        if seconds is None:
            return "Unknown"
            
        if seconds < 60:
            return f"{seconds:.1f} seconds"
        elif seconds < 3600:
            minutes = seconds / 60
            return f"{minutes:.1f} minutes"
        else:
            hours = seconds / 3600
            return f"{hours:.1f} hours"
    
    def delete_job(self, job_id: str) -> bool:
        """
        Delete a job and all its associated results.
        
        Args:
            job_id: The job identifier
            
        Returns:
            True if successful, False otherwise
        """
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Use a transaction to ensure atomicity
            cursor.execute('BEGIN TRANSACTION')
            
            # Delete from prompt_usage
            cursor.execute('DELETE FROM prompt_usage WHERE job_id = ?', (job_id,))
            
            # Delete from results
            cursor.execute('DELETE FROM results WHERE job_id = ?', (job_id,))
            
            # Delete from jobs
            cursor.execute('DELETE FROM jobs WHERE job_id = ?', (job_id,))
            
            cursor.execute('COMMIT')
            conn.close()
            
            logger.info(f"Deleted job {job_id} and associated data")
            return True
            
        except Exception as e:
            logger.error(f"Error deleting job {job_id}: {str(e)}")
            
            # Attempt to rollback transaction
            try:
                cursor.execute('ROLLBACK')
            except:
                pass
                
            return False

    def get_results(self, job_id, limit=100, offset=0):
        """
        Get results for a job with pagination.
        
        Args:
            job_id: The job ID
            limit: Maximum number of results
            offset: Number of results to skip
            
        Returns:
            List of result dictionaries
        """
        try:
            conn = self.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute(
                "SELECT * FROM results WHERE job_id = ? ORDER BY processed_at DESC LIMIT ? OFFSET ?", 
                (job_id, int(limit), int(offset))
            )
            
            rows = cursor.fetchall()
            
            # Convert rows to dicts and parse JSON fields
            results = []
            for row in rows:
                result = dict(row)
                # Parse data JSON if present
                if 'data' in result and result['data']:
                    try:
                        result['data'] = json.loads(result['data'])
                    except json.JSONDecodeError:
                        pass
                        
                # Ensure analysis_results exists for template rendering
                if 'analysis_results' not in result:
                    result['analysis_results'] = {}
                    
                results.append(result)
            
            conn.close()
            return results
        except Exception as e:
            logger.error(f"Error in get_results: {str(e)}")
            return []


# Create a global instance for easy import and use throughout the application
db = DatabaseManager()


# Test function to verify the database is working correctly
def test_database():
    """Test the database functionality."""
    try:
        # Create a database in a temporary location
        import tempfile
        test_db_path = os.path.join(tempfile.gettempdir(), "test_analysis.db")
        
        if os.path.exists(test_db_path):
            os.remove(test_db_path)
            
        test_db = DatabaseManager(test_db_path)
        
        # Create a test job - REMOVE job_id parameter from here
        job_id = test_db.create_job(
            urls=["https://example.com"],  # Add this required parameter
            prompts=["test_prompt"],  # Changed from prompt_names to prompts
            name="Test Job", 
            company_info={"name": "ACME Corp", "industry": "Technology"}
        )
        
        # Update job status
        test_db.update_job_status(
            job_id=job_id,
            status="running",
            total_urls=5,
            processed_urls=2
        )
        
        # Add a test result
        test_result = {
            "job_id": job_id,  # Include job_id in the result dict
            "url": "https://example.com",
            "title": "Example Website",
            "status": "success",
            "content_type": "html",
            "word_count": 1000,
            "processed_at": time.time(),
            "prompt_name": "test_prompt",
            "api_tokens": 500,
            "ca_target_audience": "Developers",
            "ca_quality_score": 8,
            "ca_key_themes": "Testing, Databases, Python"
        }
        
        # Pass the entire result dict to save_result
        result_id = test_db.save_result(test_result)
        
        # Get the job
        job = test_db.get_job(job_id)
        
        # Clean up
        test_db.delete_job(job_id)
        os.remove(test_db_path)
        
        print("Database test completed successfully!")
        return True
        
    except Exception as e:
        logger.error(f"Error in database test: {str(e)}")
        return False
   
if __name__ == "__main__":
    # Run the test if this module is executed directly
    test_database()
